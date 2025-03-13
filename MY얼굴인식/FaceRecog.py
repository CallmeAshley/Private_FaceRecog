import cv2
import face_recognition
import sqlite3
import numpy as np
from datetime import datetime
import openpyxl

# SQLite DB 연결
conn = sqlite3.connect('./our_faces.db')
cursor = conn.cursor()

# DB에서 얼굴 특징 벡터 불러오기
def load_face_encodings():
    cursor.execute('SELECT name, encoding FROM faces')
    rows = cursor.fetchall()
    face_encodings = []
    for row in rows:
        name, encoding_blob = row
        encoding = np.frombuffer(encoding_blob, dtype=np.float64)  # BLOB을 numpy 배열로 변환
        face_encodings.append((name, encoding))
    return face_encodings

# 얼굴 인식 후 DB 비교 함수
def recognize_and_record_face():
    # OpenCV로 웹캠 열기
    video_capture = cv2.VideoCapture(0)  # 0은 기본 웹캠

    # DB에서 얼굴 특징 벡터 로드
    known_faces = load_face_encodings()

    while True:
        ret, frame = video_capture.read()
        if not ret:
            break
        
        # 얼굴 위치와 특징 벡터 추출
        face_locations = face_recognition.face_locations(frame)
        face_encodings = face_recognition.face_encodings(frame, face_locations)

        # 얼굴이 인식되었는지 확인
        if len(face_encodings) == 0:
            # 얼굴을 인식하지 못했을 경우 "가까이 다가오세요" 메시지 띄우기
            cv2.putText(frame, "얼굴을 인식할 수 없습니다. 가까이 다가오세요.", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        else:
            for face_encoding, face_location in zip(face_encodings, face_locations):
                # DB의 얼굴 벡터와 비교
                matches = face_recognition.compare_faces([encoding for _, encoding in known_faces], face_encoding)
                name = "Unknown"  # 기본값

                # 일치하는 얼굴이 있으면
                if True in matches:
                    # 첫 번째 일치하는 얼굴의 인덱스 찾기
                    first_match_index = matches.index(True)
                    
                    # 일치하는 사람의 이름 추출
                    name = known_faces[first_match_index][0]  # known_faces에서 이름 추출

                    # 얼굴 위치 표시
                    top, right, bottom, left = face_location
                    cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)  # 얼굴 위치에 사각형 그리기
                    cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)  # 이름 표시

                    # 출퇴근 시간 기록
                    current_time = datetime.now()
                    current_date = current_time.strftime('%Y-%m-%d')
                    current_time_str = current_time.strftime('%H:%M:%S')

                    # 예시로 출근 처리
                    if 9 <= current_time.hour < 15:  # 출근 시간 범위
                        record_to_excel(name, '출근', current_date, current_time_str, '')
                    elif current_time.hour >= 18:  # 퇴근 시간 범위
                        record_to_excel(name, '퇴근', current_date, current_time_str, "9:00:00")  # 예시로 총 근무 시간 9시간
                else:
                    # DB에 얼굴이 일치하지 않으면 "얼굴을 다시 입력시키세요" 메시지 띄우기
                    cv2.putText(frame, "얼굴을 다시 입력시키세요.", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        # 화면에 결과 표시
        cv2.imshow("Face Recognition", frame)

        # 'q' 키를 눌러 종료
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()

# 엑셀에 출퇴근 기록 저장 함수
def record_to_excel(name, work_type, current_date, current_time_str, total_work_time):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook('attendance.xlsx')
    
    # 시트가 존재하지 않으면 새로 생성
    if name not in wb.sheetnames:
        sheet = wb.create_sheet(name)
        sheet.append(["날짜", "출근 시간", "퇴근 시간", "총 근무 시간"])  # 헤더 추가
    else:
        sheet = wb[name]
    
    # 출근/퇴근에 따른 기록
    if work_type == '출근':
        sheet.append([current_date, current_time_str, '', ''])  # 출근 시간 기록
    elif work_type == '퇴근':
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            if row[1].value != '' and row[2].value == '':  # 출근 시간이 있고 퇴근 시간이 비어있는 경우
                row[2].value = current_time_str  # 퇴근 시간 기록
                row[3].value = total_work_time  # 총 근무 시간 기록
                break
    
    wb.save('./attendance.xlsx')  # 엑셀 파일 저장

# 얼굴 인식 및 DB 비교 실행
recognize_and_record_face()
